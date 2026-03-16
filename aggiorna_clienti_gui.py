"""
Archivos Actualizados Automaticamente
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime
import os
import shutil
import math
import subprocess


class PizzaAnimada(tk.Canvas):

    def __init__(self, parent, size=80, **kwargs):
        super().__init__(parent, width=size, height=size + 20, highlightthickness=0, **kwargs)
        self.size = size
        self.fase = 0.0
        self.animando = False

    def iniciar(self):
        self.animando = True
        self._animar()

    def detener(self):
        self.animando = False
        self.delete("all")

    def _animar(self):
        if not self.animando:
            return
        self.delete("all")
        cx = self.size // 2
        cy = self.size // 2 + 18
        r = self.size // 2 - 4

        self.create_oval(cx - r - 4, cy - r // 3 - 2, cx + r + 4, cy + r // 3 + 8,
                         fill="#d4d4d4", outline="#bbb")
        self.create_oval(cx - r, cy - r // 2, cx + r, cy + r // 2,
                         fill="#e8a838", outline="#c4872a", width=2)
        self.create_oval(cx - r + 8, cy - r // 2 + 6, cx + r - 8, cy + r // 2 - 6,
                         fill="#d63a2a", outline="")

        for mx, my in [(-14, -6), (10, -4), (-4, 4), (16, 2), (-18, 2), (6, -10), (-8, -12)]:
            self.create_oval(cx + mx - 5, cy + my - 3, cx + mx + 5, cy + my + 3,
                             fill="#fff8e1", outline="")

        for hx, hy in [(-10, -2), (8, 4), (0, -8), (14, -4), (-6, 6)]:
            self.create_oval(cx + hx - 3, cy + hy - 2, cx + hx + 3, cy + hy + 2,
                             fill="#2e7d32", outline="")

        for i, offset_x in enumerate([-12, 0, 12]):
            fase_i = self.fase + i * 1.2
            for j in range(3):
                y_pos = cy - r // 2 - 8 - j * 8
                x_wave = math.sin(fase_i + j * 0.8) * 4
                gray = int(180 + (1 - max(0.2, 1.0 - j * 0.35)) * 75)
                self.create_line(
                    cx + offset_x + x_wave, y_pos + 4,
                    cx + offset_x - x_wave * 0.5, y_pos - 4,
                    fill=f"#{gray:02x}{gray:02x}{gray:02x}", width=2, smooth=True)

        self.fase += 0.12
        self.after(60, self._animar)


class App:

    COLONNA_ID = "Idusuario"

    def __init__(self, root):
        self.root = root
        self.root.title("Archivos Actualizados Automaticamente")
        self.root.geometry("620x440")
        self.root.resizable(False, False)

        self.file_viejo = tk.StringVar(value="Ningun archivo seleccionado")
        self.file_nuevo = tk.StringVar(value="Ningun archivo seleccionado")
        self.file_output = None
        self.ultimo_destino = None

        self._build_ui()

    def _build_ui(self):
        tk.Label(
            self.root, text="Archivos Actualizados Automaticamente",
            font=("Segoe UI", 14, "bold")
        ).pack(pady=(16, 10))

        tk.Frame(self.root, height=1, bg="#c0c0c0").pack(fill="x", padx=16)

        # Archivo Viejo
        f1 = tk.LabelFrame(self.root, text="Archivo Viejo", font=("Segoe UI", 10), padx=10, pady=10)
        f1.pack(fill="x", padx=16, pady=6)
        r1 = tk.Frame(f1)
        r1.pack(fill="x")
        tk.Label(r1, textvariable=self.file_viejo, font=("Segoe UI", 9), fg="#444",
                 anchor="w").pack(side="left", fill="x", expand=True)
        tk.Button(r1, text="Buscar...", width=12, font=("Segoe UI", 9), fg="#000000",
                  command=lambda: self._elegir(self.file_viejo)).pack(side="right", padx=(8, 0))

        # Archivo Nuevo
        f2 = tk.LabelFrame(self.root, text="Archivo Nuevo", font=("Segoe UI", 10), padx=10, pady=10)
        f2.pack(fill="x", padx=16, pady=6)
        r2 = tk.Frame(f2)
        r2.pack(fill="x")
        tk.Label(r2, textvariable=self.file_nuevo, font=("Segoe UI", 9), fg="#444",
                 anchor="w").pack(side="left", fill="x", expand=True)
        tk.Button(r2, text="Buscar...", width=12, font=("Segoe UI", 9), fg="#000000",
                  command=lambda: self._elegir(self.file_nuevo)).pack(side="right", padx=(8, 0))

        tk.Frame(self.root, height=1, bg="#c0c0c0").pack(fill="x", padx=16, pady=(10, 0))

        # Botones
        bf = tk.Frame(self.root)
        bf.pack(pady=14)

        self.btn_actualizar = tk.Button(
            bf, text="Actualizar Archivo", width=22, padx=10, pady=6,
            font=("Segoe UI", 11), fg="#000000", command=self._ejecutar_merge)
        self.btn_actualizar.pack(side="left", padx=8)

        self.btn_guardar = tk.Button(
            bf, text="Guardar Archivo", width=22, padx=10, pady=6,
            font=("Segoe UI", 11), fg="#000000", state="disabled", command=self._guardar)
        self.btn_guardar.pack(side="left", padx=8)

        # Pizza (oculta)
        self.pizza_frame = tk.Frame(self.root)

        self.pizza = PizzaAnimada(self.pizza_frame, size=80, bg=self.root.cget("bg"))
        self.pizza.pack(side="left", padx=(10, 10))

        self.lbl_pizza = tk.Label(
            self.pizza_frame,
            text="Ahora puedes invitarme\na una pizza margherita!",
            font=("Segoe UI", 11, "italic"), fg="#c44a1a", justify="left")
        self.lbl_pizza.pack(side="left", padx=(4, 0))

    # --- Logica ---

    def _elegir(self, var):
        path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Archivos Excel", "*.xlsx *.xls"), ("Todos", "*.*")])
        if path:
            var.set(path)

    def _ejecutar_merge(self):
        fv = self.file_viejo.get()
        fn = self.file_nuevo.get()

        if "Ningun archivo" in fv or "Ningun archivo" in fn:
            messagebox.showwarning("Atencion", "Selecciona ambos archivos antes de continuar.")
            return

        # Ocultar pizza si estaba visible de un uso anterior
        self.pizza.detener()
        self.pizza_frame.pack_forget()

        try:
            df_viejo = pd.read_excel(fv, dtype={self.COLONNA_ID: str})
            df_nuevo = pd.read_excel(fn, dtype={self.COLONNA_ID: str})

            for etiqueta, df in [("viejo", df_viejo), ("nuevo", df_nuevo)]:
                if self.COLONNA_ID not in df.columns:
                    cols = ", ".join(df.columns.tolist()[:15])
                    if len(df.columns) > 15:
                        cols += f"... ({len(df.columns)} columnas en total)"
                    messagebox.showerror(
                        "Error",
                        f"Columna '{self.COLONNA_ID}' no encontrada "
                        f"en el archivo {etiqueta}.\n\nColumnas: {cols}")
                    return

            ids_existentes = set(df_viejo[self.COLONNA_ID].astype(str))
            mask_nuevos = ~df_nuevo[self.COLONNA_ID].astype(str).isin(ids_existentes)
            df_agregar = df_nuevo[mask_nuevos]
            duplicados = (~mask_nuevos).sum()
            df_resultado = pd.concat([df_viejo, df_agregar], ignore_index=True)

            hoy = datetime.now().strftime("%Y-%m-%d")
            temp_dir = os.environ.get("TEMP", os.environ.get("TMP", "/tmp"))
            temp_path = os.path.join(temp_dir, f"archivo_actualizado_{hoy}.xlsx")
            df_resultado.to_excel(temp_path, index=False)

            # Evidenziare le righe nuove in verde
            fila_inicio_nuevos = len(df_viejo) + 2  # +1 header, +1 per 1-index
            num_nuevos = len(df_agregar)

            if num_nuevos > 0:
                wb = load_workbook(temp_path)
                ws = wb.active
                verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                verde_font = Font(color="006100")
                for fila in range(fila_inicio_nuevos, fila_inicio_nuevos + num_nuevos):
                    for celda in ws[fila]:
                        celda.fill = verde
                        celda.font = verde_font
                wb.save(temp_path)

            self.file_output = temp_path

            self.btn_guardar.config(state="normal")

            messagebox.showinfo(
                "Completado",
                f"Registros existentes: {len(df_viejo)}\n"
                f"Registros nuevos encontrados: {len(df_nuevo)}\n"
                f"Duplicados descartados: {duplicados}\n"
                f"Nuevos agregados: {len(df_agregar)}\n"
                f"Total final: {len(df_resultado)}\n\n"
                f"Pulsa 'Guardar Archivo' para guardar.")

        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _guardar(self):
        if not self.file_output or not os.path.exists(self.file_output):
            messagebox.showerror("Error",
                                 "No hay archivo para guardar.\nEjecuta primero la actualizacion.")
            return

        # Mostrar pizza
        self.pizza_frame.pack(fill="x", padx=20, pady=(4, 10), side="bottom")
        self.pizza.iniciar()
        self.root.update()

        # 2 segundos de pizza, luego dialogo guardar
        self.root.after(2000, self._hacer_guardado)

    def _hacer_guardado(self):
        hoy = datetime.now().strftime("%Y-%m-%d")
        nombre = f"archivo_actualizado_{hoy}.xlsx"

        # Buscar escritorio del usuario
        home = os.path.expanduser("~")
        for carpeta in ["Desktop", "Escritorio"]:
            candidato = os.path.join(home, carpeta)
            if os.path.isdir(candidato):
                inicio = candidato
                break
        else:
            inicio = home

        # Traer ventana al frente
        self.root.lift()
        self.root.focus_force()
        self.root.update()

        dest = filedialog.asksaveasfilename(
            parent=self.root,
            title="Guardar archivo actualizado",
            defaultextension=".xlsx",
            initialfile=nombre,
            initialdir=inicio,
            filetypes=[("Archivos Excel", "*.xlsx")])

        if not dest:
            return

        try:
            shutil.copy2(self.file_output, dest)
            self.ultimo_destino = dest
            carpeta = os.path.dirname(os.path.abspath(dest))

            respuesta = messagebox.askyesno(
                "Guardado",
                f"Archivo guardado correctamente en:\n\n"
                f"{dest}\n\n"
                f"Quieres abrir la carpeta?")

            if respuesta:
                if os.name == "nt":
                    subprocess.Popen(["explorer", "/select,", os.path.normpath(dest)])
                else:
                    subprocess.Popen(["xdg-open", carpeta])

        except PermissionError:
            messagebox.showerror(
                "Error de permisos",
                "No se puede guardar en esa ubicacion.\n"
                "Intenta guardar en otra carpeta (por ejemplo Documentos).")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar:\n{str(e)}")


if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()
